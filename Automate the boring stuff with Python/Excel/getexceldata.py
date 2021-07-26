import openpyxl, easygui
from openpyxl.utils import get_column_letter

file = easygui.fileopenbox(filetypes='*.xlsx')
wb = openpyxl.load_workbook(file)
print('Enter Sheet name')
sheetname = input()
sheet = wb[sheetname]

maxcolumn = get_column_letter(sheet.max_column)
maxrow = sheet.max_row
maxcell = str(maxcolumn) + str(maxrow)

for rowOfCellObjects in sheet['A1':str(maxcell)]:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--END OF ROW--')

