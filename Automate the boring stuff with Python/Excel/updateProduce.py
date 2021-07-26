#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

price_update = {'Garlic' : 3.07,
                'Celery' : 1.19,
                'Lemon'  : 1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_update:
        sheet.cell(row=rowNum, column=2).value = price_update[produceName]

wb.save('updatedProduceSales.xlsx')
