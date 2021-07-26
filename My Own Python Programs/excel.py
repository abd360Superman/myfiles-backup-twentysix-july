import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
import os
import pyinputplus as pyip
from tkinter import *
from tkinter import messagebox as mb
from tkinter import filedialog as fd

def createworkbook():
    mb.showinfo('confirmation', 'Choose new file location')
    location = fd.askdirectory()
    print('Enter name for new workbook')
    fname = input()
    fpath = os.path.join(location, fname+'.xlsx')
    wb = openpyxl.Workbook()
    wb.save(fpath)
    mb.showinfo('confirmation', 'File Created')

def createsheet():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name')
    nname = input()
    print('Enter index position for sheet, with first sheet being 0')
    indexpos = int(input())
    wb.create_sheet(index=indexpos, title=nname)
    wb.save(excelfile)
    mb.showinfo('confirmation', 'New Sheet created')

def removesheet():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter name for sheet to be deleted')
    name_to_be_del = input()
    del wb[name_to_be_del]
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Sheet deleted')

def mergecell():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name where cells will be merged')
    sname = input()
    print('Enter start cell to be merged')
    startcell = input()
    print('Enter end cell to be merged')
    endcell = input()
    sheet = wb[sname]
    mergecells = startcell + ':' + endcell
    sheet.merge_cells(mergecells)
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Cells merged')

def unmergecell():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name where cells will be unmerged')
    sname = input()
    print('Enter start cell to be unmerged')
    startcell = input()
    print('Enter end cell to be unmerged')
    endcell = input()
    sheet = wb[sname]
    unmergecells = startcell + ':' + endcell
    sheet.unmerge_cells(unmergecells)
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Cells unmerged')

def setrowheight():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name where row height will be set')
    sname = input()
    print('Enter row number for height to be set')
    rnum = int(input())
    print('Enter height for row')
    height = int(input())
    sheet = wb[sname]
    sheet.row_dimensions[rnum].height = height
    wb.save(excelfile)
    mb.showinfo('confirmation', 'New row height set')

def setcolumnwidth():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name where column width will be set')
    sname = input()
    print('Enter column alphabet where width to be set')
    halpha = input()
    print('Enter width for column')
    width = int(input())
    sheet = wb[sname]
    sheet.column_dimensions[halpha].width = width
    wb.save(excelfile)
    mb.showinfo('confirmation', 'New column width set')

def customfont():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name for custom font')
    sname = input()
    print('Enter cell for custom font to be used')
    cell = input()
    print('--START OF CUSTOM FONT--')
    print('Font name')
    foname = input()
    fosize = pyip.inputNum(prompt='Enter font size\n')
    fostyle = pyip.inputBool(prompt='Italic?\n')
    foweight = pyip.inputBool(prompt='Bold?\n')
    customFont = Font(name=foname, size=fosize, italic=fostyle, bold=foweight)
    sheet = wb[sname]
    sheet[cell].font = customFont
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Font added on chosen cell')

def changecellval():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name')
    sname = input()
    print('Enter cell')
    cell = input()
    print('Enter text to add in cell')
    text = input()
    sheet = wb[sname]
    sheet[cell].value = text
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Value added in cell')

def freezepane():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name')
    sname = input()
    print('Enter cell')
    cell = input()
    sheet = wb[sname]
    sheet.freeze_panes = cell
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Panes freezed')

def createbarchart():
    mb.showinfo('confirmation', 'Choose excel file')
    excelfile = fd.askopenfilename()
    wb = openpyxl.load_workbook(excelfile)
    print('Enter sheet name')
    sname = input()
    print('Enter start cell data range')
    startcell = input()
    print('Enter end cell data range')
    endcell = input()
    print('Enter title for bar graph')
    title = input()
    print('Which cell should the top-left corner of graph start?')
    topLeftStart = input()
    startcellDict = {'Col': '',
                     'Row': ''}
    for i in startcell:
        if i.isnumeric():
            startcellDict['Row'] += i
        else:
            startcellDict['Col'] += i
    endcellDict = {'Col': '',
                   'Row': ''}
    for j in endcell:
        if j.isnumeric():
            endcellDict['Row'] += j
        else:
            endcellDict['Col'] += j
    sheet = wb[sname]
    refObj = openpyxl.chart.Reference(sheet, min_col=int(column_index_from_string(startcellDict['Col'])), min_row=int(startcellDict['Row']), max_col=int(column_index_from_string(endcellDict['Col'])), max_row=int(endcellDict['Row']))
    sereisObj = openpyxl.chart.Series(refObj, title='First Series')
    chartObj = openpyxl.chart.BarChart()
    chartObj.title = title
    chartObj.append(sereisObj)
    sheet.add_chart(chartObj, topLeftStart)
    wb.save(excelfile)
    mb.showinfo('confirmation', 'Chart created')

tk = Tk()
tk.geometry('150x260')
tk.title(' ')
tk.resizable(0, 0)
tk.update()
Label(tk, text='Excel with Python', font=('Helvetica', 16), fg='blue').grid(row=5, column=2)

Button(tk, text='Create Workbook', command=createworkbook).grid(row=15, column=2)
Button(tk, text='Create Sheet', command=createsheet).grid(row=25, column=2)
Button(tk, text='Remove Sheet', command=removesheet).grid(row=35, column=2)
Button(tk, text='Merge Cell', command=mergecell).grid(row=45, column=2)
Button(tk, text='Unmerge Cell', command=unmergecell).grid(row=55, column=2)
Button(tk, text='Set Row Height', command=setrowheight).grid(row=65, column=2)
Button(tk, text='Set Column Width', command=setcolumnwidth).grid(row=75, column=2)
Button(tk, text='Set Custom Font', command=customfont).grid(row=85, column=2)
Button(tk, text='Change Cell Value', command=changecellval).grid(row=95, column=2)
Button(tk, text='Freeze Pane', command=freezepane).grid(row=105, column=2)
Button(tk, text='Create Bar Graph', command=createbarchart).grid(row=115, column=2)
